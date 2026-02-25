"""
AnalysisAgent: runs after financial report download. Unzips report, uses FINANCIAL_DETAILED_* only,
builds date-wise, day-of-week, slot-based, and day-slot tables (Sales, Payouts, Profitability, Orders, AOV).
Sales = Subtotal. Dollar columns (Sales, Payouts, AOV) formatted to 2 decimals. No pre/post analysis.
"""

import logging
import math
import zipfile
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple, Union

logger = logging.getLogger(__name__)

try:
    import pandas as pd
except ImportError:
    pd = None

# Slot boundaries from analysis-app slot_analysis (minutes since midnight)
SLOT_ORDER = ["Early morning", "Breakfast", "Lunch", "Afternoon", "Dinner", "Late night"]


def _find_financial_detailed_in_zip(zip_path: Path) -> Optional[str]:
    with zipfile.ZipFile(zip_path, "r") as z:
        for name in z.namelist():
            if "FINANCIAL_DETAILED" in name.upper() and name.upper().endswith(".CSV"):
                return name
    return None


def _extract_financial_detailed_csv(zip_path: Path, output_dir: Path) -> Optional[Path]:
    member = _find_financial_detailed_in_zip(zip_path)
    if not member:
        return None
    out_csv = output_dir / "financial_detailed_report.csv"
    with zipfile.ZipFile(zip_path, "r") as z:
        with z.open(member) as f:
            out_csv.write_bytes(f.read())
    logger.info("AnalysisAgent: Extracted %s", member)
    return out_csv


def _get_time_slot(time_str) -> Optional[str]:
    """Same logic as analysis-app slot_analysis.get_time_slot."""
    if pd.isna(time_str) or time_str == "":
        return None
    try:
        time_obj = pd.to_datetime(time_str, errors="coerce")
        if pd.isna(time_obj):
            return None
        total_minutes = time_obj.hour * 60 + time_obj.minute
        if total_minutes >= 0 and total_minutes < 300:
            return "Early morning"
        if total_minutes >= 300 and total_minutes < 659:
            return "Breakfast"
        if total_minutes >= 659 and total_minutes < 839:
            return "Lunch"
        if total_minutes >= 839 and total_minutes < 959:
            return "Afternoon"
        if total_minutes >= 959 and total_minutes < 1159:
            return "Dinner"
        if total_minutes >= 1159:
            return "Late night"
    except Exception:
        pass
    return None


def _resolve_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """Return (date_col, time_col, subtotal_col, payout_col, order_col). Names from analysis-app."""
    df.columns = df.columns.str.strip()
    date_col = None
    for c in ["Timestamp local date", "Timestamp Local Date", "Date", "date"]:
        if c in df.columns:
            date_col = c
            break
    time_col = None
    for c in ["Timestamp local time", "Timestamp Local Time", "Order received local time"]:
        if c in df.columns:
            time_col = c
            break
    subtotal_col = "Subtotal" if "Subtotal" in df.columns else None
    payout_col = None
    if "Net total" in df.columns:
        payout_col = "Net total"
    elif "Net total (for historical reference only)" in df.columns:
        payout_col = "Net total (for historical reference only)"
    order_col = "DoorDash order ID" if "DoorDash order ID" in df.columns else None
    return date_col, time_col, subtotal_col, payout_col, order_col


def _resolve_store_col(df: pd.DataFrame) -> Optional[str]:
    """Return Store ID column name if present."""
    df.columns = df.columns.str.strip()
    for c in ["Store ID", "Merchant store ID", "Shop ID"]:
        if c in df.columns:
            return c
    return None


def _format_dollar_columns(df: pd.DataFrame, dollar_cols: list) -> pd.DataFrame:
    """Format dollar columns as $X.XX (2 decimals)."""
    out = df.copy()
    for c in dollar_cols:
        if c in out.columns:
            out[c] = out[c].apply(lambda x: f"${float(x):,.2f}" if pd.notna(x) else "")
    return out


def _build_date_wise(df: pd.DataFrame, date_col: str, subtotal_col: str, payout_col: str, order_col: str) -> pd.DataFrame:
    """Date-wise: Sales, Payouts, Profitability (Payouts/Sales), Orders, AOV."""
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=[date_col])
    df["_date"] = df[date_col].dt.date
    df[subtotal_col] = pd.to_numeric(df[subtotal_col], errors="coerce").fillna(0)
    df[payout_col] = pd.to_numeric(df[payout_col], errors="coerce").fillna(0)
    agg = df.groupby("_date").agg(
        Sales=(subtotal_col, "sum"),
        Payouts=(payout_col, "sum"),
        Orders=(order_col, "nunique") if order_col else (subtotal_col, "count"),
    ).reset_index()
    agg.columns = ["Date", "Sales", "Payouts", "Orders"]
    agg["Profitability"] = (agg["Payouts"] / agg["Sales"].replace(0, float("nan")) * 100).round(2)
    agg["AOV"] = (agg["Sales"] / agg["Orders"].replace(0, float("nan"))).round(2)
    return agg[["Date", "Sales", "Payouts", "Profitability", "Orders", "AOV"]]


def _build_day_of_week(df: pd.DataFrame, date_col: str, subtotal_col: str, payout_col: str, order_col: str) -> pd.DataFrame:
    """Day-of-week: average (across dates) of Sales, Payouts, Profitability, Orders, AOV per weekday."""
    daily = _build_date_wise(df, date_col, subtotal_col, payout_col, order_col)
    daily["Date"] = pd.to_datetime(daily["Date"])
    daily["Day of week"] = daily["Date"].dt.day_name()
    weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    avg = daily.groupby("Day of week").agg(
        Sales=("Sales", "mean"),
        Payouts=("Payouts", "mean"),
        Profitability=("Profitability", "mean"),
        Orders=("Orders", "mean"),
        AOV=("AOV", "mean"),
    ).reset_index()
    avg["Sales"] = avg["Sales"].round(2)
    avg["Payouts"] = avg["Payouts"].round(2)
    avg["Profitability"] = avg["Profitability"].round(2)
    avg["Orders"] = avg["Orders"].round(2)
    avg["AOV"] = avg["AOV"].round(2)
    avg["Day of week"] = pd.Categorical(avg["Day of week"], categories=weekday_order, ordered=True)
    avg = avg.sort_values("Day of week").reset_index(drop=True)
    return avg[["Day of week", "Sales", "Payouts", "Profitability", "Orders", "AOV"]]


def _build_slot_based(df: pd.DataFrame, time_col: str, subtotal_col: str, payout_col: str, order_col: str) -> pd.DataFrame:
    """Slot-based: per slot Sales, Payouts, Profitability, Orders, AOV."""
    df = df.copy()
    df["_slot"] = df[time_col].apply(_get_time_slot)
    df = df.dropna(subset=["_slot"])
    df[subtotal_col] = pd.to_numeric(df[subtotal_col], errors="coerce").fillna(0)
    df[payout_col] = pd.to_numeric(df[payout_col], errors="coerce").fillna(0)
    agg = df.groupby("_slot").agg(
        Sales=(subtotal_col, "sum"),
        Payouts=(payout_col, "sum"),
        Orders=(order_col, "nunique") if order_col else (subtotal_col, "count"),
    ).reset_index()
    agg.columns = ["Slot", "Sales", "Payouts", "Orders"]
    agg["Profitability"] = (agg["Payouts"] / agg["Sales"].replace(0, float("nan")) * 100).round(2)
    agg["AOV"] = (agg["Sales"] / agg["Orders"].replace(0, float("nan"))).round(2)
    agg["Slot"] = pd.Categorical(agg["Slot"], categories=SLOT_ORDER, ordered=True)
    agg = agg.sort_values("Slot").reset_index(drop=True)
    return agg[["Slot", "Sales", "Payouts", "Profitability", "Orders", "AOV"]]


def _build_day_slot(df: pd.DataFrame, date_col: str, time_col: str, subtotal_col: str, payout_col: str, order_col: str) -> pd.DataFrame:
    """Day-Slot: Day, Slot, Sales, Payouts, Profitability, Orders, AOV, uplift, Min.Subtotal, campaign recommendation. Sorted by Day then Slot."""
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=[date_col])
    df["_day"] = df[date_col].dt.day_name()
    df["_slot"] = df[time_col].apply(_get_time_slot)
    df = df.dropna(subset=["_slot"])
    df[subtotal_col] = pd.to_numeric(df[subtotal_col], errors="coerce").fillna(0)
    df[payout_col] = pd.to_numeric(df[payout_col], errors="coerce").fillna(0)
    agg = df.groupby(["_day", "_slot"]).agg(
        Sales=(subtotal_col, "sum"),
        Payouts=(payout_col, "sum"),
        Orders=(order_col, "nunique") if order_col else (subtotal_col, "count"),
    ).reset_index()
    agg["Profitability"] = (agg["Payouts"] / agg["Sales"].replace(0, float("nan")) * 100).round(2)
    agg["AOV"] = (agg["Sales"] / agg["Orders"].replace(0, float("nan"))).round(2)
    weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    agg["Day"] = pd.Categorical(agg["_day"], categories=weekday_order, ordered=True)
    agg["Slot"] = pd.Categorical(agg["_slot"], categories=SLOT_ORDER, ordered=True)
    agg = agg.sort_values(["Day", "Slot"]).drop(columns=["_day", "_slot"]).reset_index(drop=True)
    # After AOV: uplift = AOV*1.2, Min.Subtotal = CEILING(uplift, 5), campaign recommendation
    agg["uplift"] = (agg["AOV"] * 1.2).round(2)
    agg["Min.Subtotal"] = agg["uplift"].astype(float).apply(lambda x: int(math.ceil(x / 5) * 5))
    agg["campaign recommendation"] = agg["Min.Subtotal"].apply(
        lambda m: f"All customers 15% off on min order of {m} upto Always lowest"
    )
    return agg[["Day", "Slot", "Sales", "Payouts", "Profitability", "Orders", "AOV", "uplift", "Min.Subtotal", "campaign recommendation"]]


def _build_store_slot_agg(
    df: pd.DataFrame,
    store_col: str,
    time_col: str,
    subtotal_col: str,
    payout_col: str,
    order_col: Optional[str],
) -> pd.DataFrame:
    """Aggregate by Store ID and Slot; columns Store ID, Slot, Sales, Payouts, Orders, Profitability, AOV."""
    df = df.copy()
    df["_slot"] = df[time_col].apply(_get_time_slot)
    df = df.dropna(subset=["_slot"])
    df[subtotal_col] = pd.to_numeric(df[subtotal_col], errors="coerce").fillna(0)
    df[payout_col] = pd.to_numeric(df[payout_col], errors="coerce").fillna(0)
    agg = df.groupby([store_col, "_slot"]).agg(
        Sales=(subtotal_col, "sum"),
        Payouts=(payout_col, "sum"),
        Orders=(order_col, "nunique") if order_col else (subtotal_col, "count"),
    ).reset_index()
    agg = agg.rename(columns={store_col: "Store ID", "_slot": "Slot"})
    agg["Profitability"] = (agg["Payouts"] / agg["Sales"].replace(0, float("nan")) * 100).round(2)
    agg["AOV"] = (agg["Sales"] / agg["Orders"].replace(0, float("nan"))).round(2)
    return agg


def _build_day_slot_store_agg(
    df: pd.DataFrame,
    date_col: str,
    time_col: str,
    store_col: str,
    subtotal_col: str,
    payout_col: str,
    order_col: Optional[str],
) -> pd.DataFrame:
    """Aggregate by Day-Slot and Store ID; columns Day-Slot, Store ID, Sales, Payouts, Orders, Profitability, AOV."""
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=[date_col])
    df["_day"] = df[date_col].dt.day_name()
    df["_slot"] = df[time_col].apply(_get_time_slot)
    df = df.dropna(subset=["_slot"])
    df["Day-Slot"] = df["_day"] + "-" + df["_slot"]
    df[subtotal_col] = pd.to_numeric(df[subtotal_col], errors="coerce").fillna(0)
    df[payout_col] = pd.to_numeric(df[payout_col], errors="coerce").fillna(0)
    agg = df.groupby(["Day-Slot", store_col]).agg(
        Sales=(subtotal_col, "sum"),
        Payouts=(payout_col, "sum"),
        Orders=(order_col, "nunique") if order_col else (subtotal_col, "count"),
    ).reset_index()
    agg = agg.rename(columns={store_col: "Store ID"})
    agg["Profitability"] = (agg["Payouts"] / agg["Sales"].replace(0, float("nan")) * 100).round(2)
    agg["AOV"] = (agg["Sales"] / agg["Orders"].replace(0, float("nan"))).round(2)
    return agg


def _build_store_metrics(
    df: pd.DataFrame,
    store_col: str,
    subtotal_col: str,
    payout_col: str,
    order_col: Optional[str],
) -> pd.DataFrame:
    """Per-store: Store ID, Sales, Payouts, Orders, AOV, Profitability."""
    df = df.copy()
    df[subtotal_col] = pd.to_numeric(df[subtotal_col], errors="coerce").fillna(0)
    df[payout_col] = pd.to_numeric(df[payout_col], errors="coerce").fillna(0)
    agg = df.groupby(store_col).agg(
        Sales=(subtotal_col, "sum"),
        Payouts=(payout_col, "sum"),
        Orders=(order_col, "nunique") if order_col else (subtotal_col, "count"),
    ).reset_index()
    agg = agg.rename(columns={store_col: "Store ID"})
    agg["Profitability"] = (agg["Payouts"] / agg["Sales"].replace(0, float("nan")) * 100).round(2)
    agg["AOV"] = (agg["Sales"] / agg["Orders"].replace(0, float("nan"))).round(2)
    return agg[["Store ID", "Sales", "Payouts", "Profitability", "Orders", "AOV"]]


def _build_campaign_recommendations(store_metrics: pd.DataFrame) -> pd.DataFrame:
    """
    Campaign recommendations per store from AOV.
    B = MROUND(AOV, 5), A = 20% if B > AOV else 15%.
    Rec1: New customers {A}% off on min order of {B} upto Always lowest.
    C = CEILING(AOV*1.2, 5). Rec2: All customers 15% off on min order of {C} upto Always lowest.
    """
    if store_metrics.empty or "AOV" not in store_metrics.columns:
        return pd.DataFrame()
    out = store_metrics[["Store ID", "AOV"]].copy()
    aov = out["AOV"].astype(float)
    # B = MROUND(AOV, 5)
    B = (aov / 5).round() * 5
    B = B.clip(lower=5)  # at least 5
    # A = 20 if B > AOV else 15
    A = (20 * (B > aov) + 15 * (B <= aov)).astype(int)
    # C = CEILING(AOV*1.2, 5)
    C = aov.apply(lambda x: math.ceil((float(x) * 1.2) / 5) * 5)
    C = C.clip(lower=5)
    out["Min order (new cust) B"] = B
    out["Discount % (new cust) A"] = A
    out["Recommendation 1"] = (
        "New customers " + A.astype(str) + "% off on min order of $" + B.astype(int).astype(str) + " upto Always lowest"
    )
    out["Min order (all cust) C"] = C
    out["Recommendation 2"] = (
        "All customers 15% off on min order of $" + C.astype(int).astype(str) + " upto Always lowest"
    )
    return out[
        [
            "Store ID",
            "AOV",
            "Min order (new cust) B",
            "Discount % (new cust) A",
            "Recommendation 1",
            "Min order (all cust) C",
            "Recommendation 2",
        ]
    ]


DOLLAR_COLS = ["Sales", "Payouts", "AOV"]


def _write_excel(
    output_dir: Path,
    date_wise: pd.DataFrame,
    day_of_week: pd.DataFrame,
    slot: pd.DataFrame,
    day_slot: pd.DataFrame,
    store_wise: pd.DataFrame,
    campaign_recs: pd.DataFrame,
    store_slot_pivots: Optional[list] = None,
    day_slot_store_pivots: Optional[list] = None,
    operator_name: Optional[str] = None,
    day_slot_per_store: Optional[List[Tuple[str, pd.DataFrame]]] = None,
) -> Path:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    wb.remove(wb.active)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = (operator_name.strip() if operator_name and isinstance(operator_name, str) else None)
    filename = f"{tag}_financial_analysis_{timestamp}.xlsx" if tag else f"financial_analysis_{timestamp}.xlsx"
    filepath = output_dir / filename

    def add_sheet(ws, df, title: str):
        if df is None or df.empty:
            return
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = Font(bold=True)

    ws1 = wb.create_sheet("Date-wise")
    add_sheet(ws1, date_wise, "Date-wise: Sales, Payouts, Profitability, Orders, AOV")
    if store_wise is not None and not store_wise.empty:
        ws_store = wb.create_sheet("Store-wise")
        add_sheet(ws_store, store_wise, "Store-wise: Sales, Payouts, Profitability, Orders, AOV (by Store ID)")
    ws2 = wb.create_sheet("Day of week")
    add_sheet(ws2, day_of_week, "Day-of-week averages: Sales, Payouts, Profitability, Orders, AOV")
    ws3 = wb.create_sheet("Slot-based")
    add_sheet(ws3, slot, "Slot-based: Sales, Payouts, Profitability, Orders, AOV")
    if day_slot is not None and not day_slot.empty:
        ws4 = wb.create_sheet("Day-Slot")
        add_sheet(ws4, day_slot, "Day-Slot: Day, Slot, Sales, Payouts, Profitability, Orders, AOV, uplift, Min.Subtotal, campaign recommendation")
    if day_slot_per_store:
        for sheet_name, tbl in day_slot_per_store:
            if tbl is not None and not tbl.empty:
                ws = wb.create_sheet(sheet_name[:31])
                add_sheet(ws, tbl, f"Day-Slot: {sheet_name}")

    for sheet_name, pivot_df in (store_slot_pivots or []) + (day_slot_store_pivots or []):
        if pivot_df is not None and not pivot_df.empty:
            ws = wb.create_sheet(sheet_name[:31])
            add_sheet(ws, pivot_df, sheet_name)

    wb.save(filepath)
    logger.info("AnalysisAgent: Wrote %s", filepath.name)
    return filepath


def run(
    zip_path: Path,
    output_dir: Path,
    report_start_date: str,
    report_end_date: str,
    excluded_dates: Optional[list] = None,
    operator_name: Optional[str] = None,
    write_file: bool = True,
) -> Union[Optional[Path], Optional[List[Tuple[str, pd.DataFrame]]]]:
    """
    Load FINANCIAL_DETAILED_* from zip, build date-wise / day-of-week / slot-based tables.
    If write_file=True, writes financial_analysis_<timestamp>.xlsx and returns path.
    If write_file=False, returns list of (sheet_name, DataFrame) for combined report.
    """
    if pd is None:
        raise RuntimeError("pandas is required for AnalysisAgent. Install with: pip install pandas")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    zip_path = Path(zip_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    extracted_csv = _extract_financial_detailed_csv(zip_path, output_dir)
    if not extracted_csv or not extracted_csv.is_file():
        logger.warning("AnalysisAgent: No FINANCIAL_DETAILED_* in zip")
        return None

    df = pd.read_csv(extracted_csv)
    date_col, time_col, subtotal_col, payout_col, order_col = _resolve_columns(df)
    if not all([date_col, subtotal_col, payout_col]):
        logger.warning("AnalysisAgent: Missing required columns (date, Subtotal, Net total)")
        return None

    store_col = _resolve_store_col(df)
    date_wise = _build_date_wise(df, date_col, subtotal_col, payout_col, order_col or subtotal_col)
    day_of_week = _build_day_of_week(df, date_col, subtotal_col, payout_col, order_col or subtotal_col)
    slot_table = _build_slot_based(df, time_col, subtotal_col, payout_col, order_col) if time_col else pd.DataFrame()
    day_slot_table = _build_day_slot(df, date_col, time_col, subtotal_col, payout_col, order_col) if time_col else pd.DataFrame()
    day_slot_per_store: List[Tuple[str, pd.DataFrame]] = []
    if store_col and time_col and not day_slot_table.empty:
        for store_id in df[store_col].dropna().unique():
            store_df = df[df[store_col] == store_id]
            tbl = _build_day_slot(store_df, date_col, time_col, subtotal_col, payout_col, order_col)
            if not tbl.empty:
                tbl = _format_dollar_columns(tbl, [c for c in DOLLAR_COLS + ["uplift"] if c in tbl.columns])
                sheet_name = f"Day-Slot - {store_id}"[:31]
                day_slot_per_store.append((sheet_name, tbl))
    store_metrics = _build_store_metrics(df, store_col, subtotal_col, payout_col, order_col) if store_col else pd.DataFrame()
    store_wise = store_metrics.copy()
    campaign_recs = _build_campaign_recommendations(store_metrics) if not store_metrics.empty else pd.DataFrame()
    if not campaign_recs.empty:
        campaign_recs = _format_dollar_columns(campaign_recs, [c for c in ["AOV", "Min order (new cust) B", "Min order (all cust) C"] if c in campaign_recs.columns])

    store_slot_pivots = []
    day_slot_store_pivots = []
    if store_col and time_col:
        store_slot_agg = _build_store_slot_agg(df, store_col, time_col, subtotal_col, payout_col, order_col)
        if not store_slot_agg.empty:
            for metric in ["AOV", "Profitability", "Sales", "Payouts", "Orders"]:
                if metric in store_slot_agg.columns:
                    pt = store_slot_agg.pivot(index="Store ID", columns="Slot", values=metric)
                    pt = pt.reindex(columns=[s for s in SLOT_ORDER if s in pt.columns])
                    pt = pt.reset_index()
                    if metric in DOLLAR_COLS:
                        dollar_cols = [c for c in pt.columns if c != "Store ID"]
                        pt = _format_dollar_columns(pt, dollar_cols)
                    store_slot_pivots.append((f"Store-Slot {metric}", pt))
        day_slot_store_agg = _build_day_slot_store_agg(df, date_col, time_col, store_col, subtotal_col, payout_col, order_col)
        if not day_slot_store_agg.empty:
            for metric in ["AOV", "Profitability", "Sales", "Payouts", "Orders"]:
                if metric in day_slot_store_agg.columns:
                    pt = day_slot_store_agg.pivot(index="Day-Slot", columns="Store ID", values=metric)
                    pt = pt.reset_index()
                    if metric in DOLLAR_COLS:
                        dollar_cols = [c for c in pt.columns if c != "Day-Slot"]
                        pt = _format_dollar_columns(pt, dollar_cols)
                    day_slot_store_pivots.append((f"DaySlot-Store {metric}", pt))

    date_wise = _format_dollar_columns(date_wise, [c for c in DOLLAR_COLS if c in date_wise.columns]) if not date_wise.empty else date_wise
    store_wise = _format_dollar_columns(store_wise, [c for c in DOLLAR_COLS if c in store_wise.columns]) if not store_wise.empty else store_wise
    day_of_week = _format_dollar_columns(day_of_week, [c for c in DOLLAR_COLS if c in day_of_week.columns]) if not day_of_week.empty else day_of_week
    slot_table = _format_dollar_columns(slot_table, [c for c in DOLLAR_COLS if c in slot_table.columns]) if not slot_table.empty else slot_table
    day_slot_table = _format_dollar_columns(day_slot_table, [c for c in DOLLAR_COLS + ["uplift"] if c in day_slot_table.columns]) if not day_slot_table.empty else day_slot_table

    sheets_list: List[Tuple[str, pd.DataFrame]] = [
        ("Date-wise", date_wise),
        ("Store-wise", store_wise),
        ("Day of week", day_of_week),
        ("Slot-based", slot_table),
    ]
    # Always include combined Day-Slot sheet when we have data (for combined report and standalone)
    if not day_slot_table.empty:
        sheets_list.append(("Day-Slot", day_slot_table))
    if day_slot_per_store:
        for sheet_name, tbl in day_slot_per_store:
            sheets_list.append((sheet_name, tbl))
    for name, pt_df in (store_slot_pivots or []) + (day_slot_store_pivots or []):
        if pt_df is not None and not pt_df.empty:
            sheets_list.append((name[:31], pt_df))

    if not write_file:
        return sheets_list
    try:
        return _write_excel(
            output_dir, date_wise, day_of_week, slot_table, day_slot_table, store_wise, campaign_recs,
            store_slot_pivots, day_slot_store_pivots, operator_name, day_slot_per_store=day_slot_per_store if day_slot_per_store else None,
        )
    except Exception as e:
        logger.warning("AnalysisAgent: Failed to write Excel: %s", e)
        return None
