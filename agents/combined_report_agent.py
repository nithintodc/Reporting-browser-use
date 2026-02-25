"""
CombinedReportAgent: build one Excel workbook with all financial + marketing analysis sheets.
Can write from in-memory sheet data (DataFrames) or from existing xlsx files.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple, Union

logger = logging.getLogger(__name__)

try:
    import pandas as pd
except ImportError:
    pd = None


def _copy_sheet_from_book(src_wb, sheet_name, dest_wb, new_name=None):
    """Copy a sheet from src_wb to dest_wb (values only)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")
    if sheet_name not in src_wb.sheetnames:
        return
    src_ws = src_wb[sheet_name]
    name = (new_name or sheet_name)[:31]
    if name in dest_wb.sheetnames:
        base, n = name, 1
        while f"{base}_{n}"[:31] in dest_wb.sheetnames:
            n += 1
        name = f"{base}_{n}"[:31]
    dest_ws = dest_wb.create_sheet(name)
    for row in src_ws.iter_rows():
        for cell in row:
            dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)


def write_combined_report(
    financial_xlsx_path: Optional[Path] = None,
    marketing_xlsx_path: Optional[Path] = None,
    output_dir: Optional[Path] = None,
    output_filename: Optional[str] = None,
) -> Optional[Path]:
    """
    Create one Excel workbook with all sheets from financial and marketing workbooks.
    Financial sheets first (with "Financial - " prefix on sheet names if needed to avoid clashes),
    then marketing sheets. Saves to output_dir/combined_analysis_{timestamp}.xlsx unless output_filename set.
    Returns path to the combined file or None.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl required for combined report")
        return None

    output_dir = Path(output_dir) if output_dir else Path("downloads")
    output_dir.mkdir(parents=True, exist_ok=True)
    if not output_filename:
        output_filename = f"combined_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = output_dir / output_filename

    wb_out = openpyxl.Workbook()
    # Remove default sheet after we add first real sheet
    default_sheet = wb_out.active
    sheet_count = 0

    for label, xlsx_path in [("Financial", financial_xlsx_path), ("Marketing", marketing_xlsx_path)]:
        if not xlsx_path or not Path(xlsx_path).is_file():
            continue
        try:
            wb_src = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
        except Exception as e:
            logger.warning("Could not open %s: %s", xlsx_path, e)
            continue
        for name in wb_src.sheetnames:
            safe_name = name[:31]
            if safe_name in wb_out.sheetnames:
                safe_name = f"{label}-{name}"[:31]
            _copy_sheet_from_book(wb_src, name, wb_out, safe_name)
            sheet_count += 1
        wb_src.close()

    if sheet_count == 0:
        wb_out.close()
        logger.warning("CombinedReportAgent: No sheets to write")
        return None

    wb_out.remove(default_sheet)
    wb_out.save(out_path)
    logger.info("CombinedReportAgent: Wrote %s (%s sheets)", out_path.name, sheet_count)
    return out_path


def _add_sheet_from_df(wb, sheet_name: str, df, title: str = None):
    """Add a sheet to openpyxl workbook from a pandas DataFrame."""
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    name = (sheet_name or "Sheet")[:31]
    if name in wb.sheetnames:
        base, n = name, 1
        while f"{base}_{n}"[:31] in wb.sheetnames:
            n += 1
        name = f"{base}_{n}"[:31]
    ws = wb.create_sheet(name)
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        start_row = 3
    else:
        start_row = 1
    if df is not None and not (getattr(df, "empty", True)):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True)


def write_combined_from_sheets(
    financial_sheets: Optional[List[Tuple[str, object]]] = None,
    marketing_sheets: Optional[List[Tuple[str, object]]] = None,
    output_dir: Optional[Path] = None,
    output_filename: Optional[str] = None,
) -> Optional[Path]:
    """
    Build one workbook from list of (sheet_name, DataFrame) for financial and marketing.
    Saves to output_dir/combined_analysis_{timestamp}.xlsx. Returns path or None.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl required for combined report")
        return None
    financial_sheets = financial_sheets or []
    marketing_sheets = marketing_sheets or []
    if not financial_sheets and not marketing_sheets:
        logger.warning("CombinedReportAgent: No sheets to write")
        return None

    output_dir = Path(output_dir) if output_dir else Path("downloads")
    output_dir.mkdir(parents=True, exist_ok=True)
    if not output_filename:
        output_filename = f"combined_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = output_dir / output_filename

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    sheet_count = 0
    for name, df in financial_sheets:
        if df is not None and not (getattr(df, "empty", True)):
            _add_sheet_from_df(wb, name, df, name)
            sheet_count += 1
    for name, df in marketing_sheets:
        if df is not None and not (getattr(df, "empty", True)):
            safe = name[:31]
            if safe in wb.sheetnames:
                safe = f"Marketing-{name}"[:31]
            _add_sheet_from_df(wb, safe, df, name)
            sheet_count += 1
    if sheet_count == 0:
        wb.close()
        return None
    wb.remove(default_sheet)
    wb.save(out_path)
    logger.info("CombinedReportAgent: Wrote %s (%s sheets)", out_path.name, sheet_count)
    return out_path


def run(
    financial_xlsx_path: Optional[Path] = None,
    marketing_xlsx_path: Optional[Path] = None,
    financial_sheets: Optional[List[Tuple[str, object]]] = None,
    marketing_sheets: Optional[List[Tuple[str, object]]] = None,
    output_dir: Optional[Path] = None,
) -> Optional[Path]:
    """Write combined workbook from either xlsx paths or in-memory sheet lists. Returns path or None."""
    if financial_sheets is not None or marketing_sheets is not None:
        return write_combined_from_sheets(
            financial_sheets=financial_sheets,
            marketing_sheets=marketing_sheets,
            output_dir=output_dir,
        )
    return write_combined_report(
        financial_xlsx_path=financial_xlsx_path,
        marketing_xlsx_path=marketing_xlsx_path,
        output_dir=output_dir,
    )
